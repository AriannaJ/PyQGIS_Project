from openpyxl import *

# Locate and Read Excel File from Desktop
try:
    path = input("Path to Excel File(Wallace vs. Wiedower): ")
    book = load_workbook(path)
except:
    path = input("File Not Valid. Check Spelling and Try Again. Make sure to add the '.xlsx' extention to end "
                 "of file path: ")
    book = load_workbook(path)

sheet = book["Election Night Raw Data"]

# Create New Excel File to Write Output to
finished_product = Workbook()
finished_sheet = finished_product.active

# Copy 1st Column (Precinct Names) from Original Sheet
for i in range(1,10):
    finished_sheet.cell(row=i,column=1).value = sheet.cell(row=i,column=1).value

# Generate Remaining Column Headers for Output File
finished_sheet['B1'] = "Jonathan Wallace"
finished_sheet['C1'] = "Marcus Wiedower"
finished_sheet['D1'] = "Total Votes"
finished_sheet['E1'] = "Winner"
finished_sheet['F1'] = "Difference"
finished_sheet['G1'] = "DEM %"
finished_sheet['H1'] = "REP %"
finished_sheet['I1'] = "Bench (D)"
finished_sheet['J1'] = "Bench % (D)"
finished_sheet['K1'] = "Bench (R)"
finished_sheet['L1'] = "Bench % (R)"

# This function extracts voting data for each candidate, sums the data,
# and writes it to Output
def print_results(x, y, z):
    sum_votes = 0

    for j in range(2, 10):
        for i in range(x, y):
            num_vot = int(sheet.cell(row=j, column=i).value)
            sum_votes = sum_votes + num_vot
            finished_sheet.cell(row=j, column=z).value = sum_votes

        sum_votes = 0

# This function calculates the number of votes for both candidates
# in all Precincts
def total_votes():
    total_votes_both = 0
    for j in range(2,10):
        for i in range(2,4):
            total_votes_both = total_votes_both + int(finished_sheet.cell(row=j,column=i).value)
        print("total " + str(total_votes_both))

        # Write Info to Output
        finished_sheet.cell(row=j, column=(i + 1)).value = total_votes_both
        total_votes_both = 0


# This function determines the party that won each precinct,
# the difference in votes between the winning party and the losing party,
# and writes it to output
def determine_winner_diff():
        for j in range(2, 10):
            for i in range(2, 3):
                if int(finished_sheet.cell(row=j, column=i).value) < int(finished_sheet.cell(row=j, column=(i+1)).value):
                    finished_sheet.cell(row=j, column=5).value = "REP"
                    print("REP")
                else:
                    finished_sheet.cell(row=j, column=5).value = "DEM"
                    print("DEM")

                # Calculate 'Difference in Vote' and Write to Output
                finished_sheet.cell(row=j,column=6).value = int(finished_sheet.cell(row=j, column=i).value) - int(finished_sheet.cell(row=j, column=(i + 1)).value)

# This function determines percentage of vote for given party,x
def pctg(x):
    for i in range(2,10):
        if x == "DEM":
            finished_sheet.cell(row=i, column=7).value = (((int(finished_sheet.cell(row=i, column=2).value))
                                                         / finished_sheet.cell(row=i, column=4).value)) * 100
        elif x == "REP":
            finished_sheet.cell(row=i,column=8).value = (((int(finished_sheet.cell(row=i, column=3).value))
                                                        / finished_sheet.cell(row=i, column=4).value)) * 100
        else:
            print("Value entered is wrong")


def ask_for_athens_benchmark(x):
    for i in range(2,10):
        if x == "DEM":
            benchmark = int(input("Enter Benchmark for Athens " + finished_sheet.cell(row=i, column=1).value + ":"))
            finished_sheet.cell(row=i, column=9).value = benchmark
            print("Benchmark Progress: " + str(finished_sheet.cell(row=i, column=2).value))
            bm = (finished_sheet.cell(row=i, column=2).value/benchmark) * 100
            finished_sheet.cell(row=i, column=10).value = bm
            print("Benchmark %: " + str(bm))
        elif x == "REP":
            benchmark = int(input("Enter Benchmark for Athens " + finished_sheet.cell(row=i, column=1).value + ":"))
            finished_sheet.cell(row=i, column=11).value = benchmark
            print("Benchmark Progress: " + str(finished_sheet.cell(row=i, column=3).value))
            bm = (finished_sheet.cell(row=i, column=2).value / benchmark) * 100
            finished_sheet.cell(row=i, column=12).value = bm
            print("Benchmark %: " + str(bm))


print_results(3, 6, 2)
print_results(6, 9, 3)
total_votes()
determine_winner_diff()
pctg("DEM")
pctg("REP")
# ask_for_athens_benchmark("DEM") . This is a function I commented out to save time on quick demonstrations

finished_product.save("ENight.xlsx")
